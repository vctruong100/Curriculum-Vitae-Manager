"""
Excel (.xlsx) parser for master study lists.
Handles Column A (hierarchy), Column B (full description), Column C (masked description).
"""

import re
from pathlib import Path
from typing import List, Optional, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

import logging

from models import Study, ResearchExperience, Phase, Subcategory
from normalizer import (
    normalize_phase, normalize_for_display, extract_protocol,
    parse_sponsor_protocol, is_phase_heading, validate_year
)

SEVEN_COL_HEADERS = [
    "Phase",
    "Subcategory",
    "Year",
    "Sponsor",
    "Protocol",
    "Masked Description",
    "Full Description",
]


def parse_master_xlsx(file_path: Path) -> List[Study]:
    """
    Parse a master .xlsx file into a list of Study objects.
    
    Expected format:
    - Column A: Hierarchy stream (Phase row -> Subcategory row -> Year rows)
    - Column B: DescriptionFull (includes protocol and treatment)
    - Column C: DescriptionMasked (no protocol; treatments replaced by XXX/XXXXXX)
    
    Returns normalized list of Study objects.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    studies = []
    current_phase = ""
    current_subcategory = ""
    
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or all(cell is None for cell in row):
            continue
        
        col_a = str(row[0]).strip() if row[0] else ""
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        col_c = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        
        if not col_a:
            continue
        
        # Check if this is a phase heading
        phase_name = is_phase_heading(col_a)
        if phase_name:
            current_phase = normalize_phase(phase_name)
            continue
        
        # Check if this is a year line (study entry)
        year_match = re.match(r'^(\d{4})(?:[\t\s]|$)', col_a)
        if year_match:
            year = int(year_match.group(1))
            
            # Parse Column B to extract sponsor, protocol, description_full
            sponsor, protocol, desc_full = parse_column_b(col_b)
            
            # Column C is the masked description - parse it to extract just the description
            # (Column C format is same as Column B: "Sponsor: Description" but without protocol)
            if col_c:
                _, _, desc_masked = parse_column_b(col_c)
                if not desc_masked:
                    desc_masked = col_c  # Fallback if no colon found
            else:
                desc_masked = desc_full
            
            # If no description in column B, check if column A has more content
            if not desc_full and len(col_a) > 4:
                rest = col_a[year_match.end():].strip()
                sponsor, protocol, desc_full = parse_column_b(rest)
                if not desc_masked:
                    desc_masked = desc_full
            
            study = Study(
                phase=current_phase,
                subcategory=current_subcategory,
                year=year,
                sponsor=sponsor,
                protocol=protocol,
                description_full=desc_full,
                description_masked=desc_masked,
            )
            studies.append(study)
        else:
            # This is a subcategory heading
            current_subcategory = normalize_for_display(col_a)
    
    wb.close()
    return studies


def parse_column_b(text: str) -> Tuple[str, str, str]:
    """
    Parse Column B to extract Sponsor, Protocol, and Description.
    
    Format: {Sponsor}{[ SPACE ]{Protocol}}: {Description}
    
    Returns: (sponsor, protocol, description)
    """
    if not text:
        return "", "", ""
    
    text = normalize_for_display(text)
    
    # Find the first colon to split sponsor/protocol from description
    colon_idx = text.find(':')
    
    if colon_idx == -1:
        # No colon - try to extract what we can
        sponsor, protocol = parse_sponsor_protocol(text)
        return sponsor, protocol, ""
    
    sponsor_protocol_part = text[:colon_idx].strip()
    description = text[colon_idx + 1:].strip()
    
    # Parse sponsor and protocol
    sponsor, protocol = parse_sponsor_protocol(sponsor_protocol_part)
    
    return sponsor, protocol, description


def studies_to_research_experience(studies: List[Study]) -> ResearchExperience:
    """Convert a list of studies to a ResearchExperience structure."""
    re_struct = ResearchExperience()
    
    for study in studies:
        phase = re_struct.get_or_create_phase(study.phase)
        subcategory = phase.get_or_create_subcategory(study.subcategory)
        subcategory.studies.append(study)
    
    re_struct.sort_all()
    return re_struct


def export_studies_to_xlsx(
    studies: List[Study],
    output_path: Path,
    include_hierarchy: bool = True,
    custom_order: Optional[List[str]] = None
) -> None:
    """
    Export studies to an Excel file in the master format.
    
    Column A: Hierarchy (Phase/Subcategory/Year)
    Column B: Full description with protocol
    Column C: Masked description
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Studies"
    
    # No headers - start directly with data
    row = 1
    current_phase = ""
    current_subcategory = ""
    
    # Sort studies based on custom order if provided, otherwise use default
    if custom_order:
        # Create order lookup
        order_lookup = {key: idx for idx, key in enumerate(custom_order)}
        default_order = len(custom_order)
        
        sorted_studies = sorted(studies, key=lambda s: (
            order_lookup.get(f"{s.phase} > {s.subcategory}", default_order),
            -s.year,
            s.sponsor.lower(),
            s.protocol.lower()
        ))
    else:
        # Default sorting
        sorted_studies = sorted(studies, key=lambda s: (
            0 if "phase i" in s.phase.lower() and "ii" not in s.phase.lower() else 1,
            s.subcategory.lower(),
            -s.year,
            s.sponsor.lower(),
            s.protocol.lower()
        ))
    
    for study in sorted_studies:
        # Write phase heading if changed
        if include_hierarchy and study.phase != current_phase:
            current_phase = study.phase
            ws.cell(row=row, column=1, value=current_phase)
            row += 1
            current_subcategory = ""  # Reset subcategory when phase changes
        
        # Write subcategory heading if changed
        if include_hierarchy and study.subcategory != current_subcategory:
            current_subcategory = study.subcategory
            ws.cell(row=row, column=1, value=current_subcategory)
            row += 1
        
        # Write study row
        # Column A: Year
        ws.cell(row=row, column=1, value=study.year)
        
        # Column B: Full description
        if study.protocol:
            full_text = f"{study.sponsor} {study.protocol}: {study.description_full}"
        else:
            full_text = f"{study.sponsor}: {study.description_full}"
        ws.cell(row=row, column=2, value=full_text)
        
        # Column C: Masked description
        masked_text = f"{study.sponsor}: {study.description_masked}"
        ws.cell(row=row, column=3, value=masked_text)
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 80
    
    wb.save(output_path)
    wb.close()


def detect_xlsx_format(file_path: Path) -> str:
    """Detect whether an .xlsx file uses the 7-column or legacy 3-column format.

    Returns:
        '7col' if the first row matches SEVEN_COL_HEADERS exactly.
        '3col' if the file appears to use the legacy hierarchy format.
        'unknown' otherwise.
    """
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        first_row = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            first_row = [str(c).strip() if c is not None else "" for c in row]
            break
        wb.close()
        if len(first_row) >= 7:
            if first_row[:7] == SEVEN_COL_HEADERS:
                return "7col"
        return "3col"
    except Exception:
        return "unknown"


def parse_master_xlsx_seven_col(file_path: Path) -> List[Study]:
    """Parse a 7-column .xlsx file into a list of Study objects.

    Expected columns (row 1 header):
        Phase | Subcategory | Year | Sponsor | Protocol | Masked Description | Full Description

    Year must be an integer or numeric-convertible value.
    Returns a list of Study objects.
    Raises ValueError on invalid headers or data.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        raise ValueError("File is empty")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    if len(header) < 7:
        raise ValueError(
            f"Expected 7 columns ({', '.join(SEVEN_COL_HEADERS)}), "
            f"got {len(header)}. If using legacy 3-column format, "
            f"please re-export with the 7-column schema."
        )
    for idx, expected in enumerate(SEVEN_COL_HEADERS):
        if header[idx] != expected:
            raise ValueError(
                f"Column {idx + 1} header must be '{expected}', "
                f"got '{header[idx]}'. Expected headers: {SEVEN_COL_HEADERS}"
            )

    studies = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(c is None for c in row):
            continue
        cells = list(row) + [None] * max(0, 7 - len(row))

        phase = str(cells[0]).strip() if cells[0] else ""
        subcategory = str(cells[1]).strip() if cells[1] else ""
        year_raw = cells[2]
        sponsor = str(cells[3]).strip() if cells[3] else ""
        protocol = str(cells[4]).strip() if cells[4] else ""
        desc_masked = str(cells[5]).strip() if cells[5] else ""
        desc_full = str(cells[6]).strip() if cells[6] else ""

        if not phase and not sponsor:
            continue

        try:
            year_val = int(float(str(year_raw))) if year_raw is not None else 0
        except (ValueError, TypeError):
            raise ValueError(
                f"Row {row_num}: Year must be numeric, got {year_raw!r}"
            )
        if year_val != 0 and not (1900 <= year_val <= 2100):
            raise ValueError(
                f"Row {row_num}: Year must be 1900-2100 or 0, got {year_val}"
            )

        phase = normalize_phase(phase)

        study = Study(
            phase=phase,
            subcategory=subcategory,
            year=year_val,
            sponsor=sponsor,
            protocol=protocol,
            description_full=desc_full,
            description_masked=desc_masked,
        )
        studies.append(study)

    logging.info(
        "[ExcelParser] Parsed %d studies from 7-column file '%s'",
        len(studies),
        file_path.name,
    )
    return studies


def export_studies_to_xlsx_seven_col(
    studies: List[Study],
    output_path: Path,
    custom_order: Optional[List[str]] = None,
) -> None:
    """Export studies to a 7-column .xlsx file.

    Columns: Phase | Subcategory | Year | Sponsor | Protocol | Masked Description | Full Description
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Studies"

    for col_idx, header in enumerate(SEVEN_COL_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    if custom_order:
        order_lookup = {key: idx for idx, key in enumerate(custom_order)}
        default_order = len(custom_order)
        sorted_studies = sorted(studies, key=lambda s: (
            order_lookup.get(f"{s.phase} > {s.subcategory}", default_order),
            -s.year,
            s.sponsor.lower(),
            s.protocol.lower(),
        ))
    else:
        sorted_studies = sorted(studies, key=lambda s: (
            0 if "phase i" in s.phase.lower() and "ii" not in s.phase.lower() else 1,
            s.subcategory.lower(),
            -s.year,
            s.sponsor.lower(),
            s.protocol.lower(),
        ))

    for row_idx, study in enumerate(sorted_studies, start=2):
        ws.cell(row=row_idx, column=1, value=study.phase)
        ws.cell(row=row_idx, column=2, value=study.subcategory)
        ws.cell(row=row_idx, column=3, value=study.year)
        ws.cell(row=row_idx, column=4, value=study.sponsor)
        ws.cell(row=row_idx, column=5, value=study.protocol)
        ws.cell(row=row_idx, column=6, value=study.description_masked)
        ws.cell(row=row_idx, column=7, value=study.description_full)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 60
    ws.column_dimensions['G'].width = 60

    wb.save(output_path)
    wb.close()
    logging.info(
        "[ExcelParser] Exported %d studies to 7-column file '%s'",
        len(sorted_studies),
        output_path.name,
    )


def validate_master_xlsx(file_path: Path) -> Tuple[bool, str]:
    """
    Validate a master .xlsx file format.
    
    Returns: (is_valid, error_message)
    """
    try:
        if not file_path.exists():
            return False, f"File not found: {file_path}"
        
        if not file_path.suffix.lower() == '.xlsx':
            return False, "File must be a .xlsx file"
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Check if there's any data
        has_data = False
        has_phase = False
        has_year = False
        
        for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
            if not row or all(cell is None for cell in row):
                continue
            
            has_data = True
            col_a = str(row[0]).strip() if row[0] else ""
            
            if is_phase_heading(col_a):
                has_phase = True
            
            if re.match(r'^\d{4}(?:[\t\s]|$)', col_a):
                has_year = True
        
        wb.close()
        
        if not has_data:
            return False, "File appears to be empty"
        
        if not has_phase:
            return False, "No phase headings found (e.g., 'Phase I', 'Phase II-IV')"
        
        if not has_year:
            return False, "No study entries found (lines starting with year)"
        
        return True, ""
        
    except Exception as e:
        return False, f"Error reading file: {str(e)}"
