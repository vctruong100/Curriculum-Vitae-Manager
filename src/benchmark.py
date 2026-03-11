"""
Micro-benchmark script for the CV Research Experience Manager.

Measures performance of key operations at scale:
- Master list parsing (1k, 5k, 10k studies)
- Fuzzy matching throughput
- Normalisation throughput
- SQLite bulk insert/read

Target latencies on a modest laptop (i5, 8 GB RAM, SSD):
  1k studies:  parse < 1 s, match < 2 s, bulk insert < 1 s
  5k studies:  parse < 3 s, match < 8 s, bulk insert < 3 s
  10k studies: parse < 6 s, match < 20 s, bulk insert < 5 s

Usage:
    python benchmark.py [--count 1000]
"""

import argparse
import sys
import time
import tempfile
import sqlite3
from pathlib import Path

# Ensure app dir on path
app_dir = Path(__file__).parent.resolve()
if str(app_dir) not in sys.path:
    sys.path.insert(0, str(app_dir))

from openpyxl import Workbook
from models import Study
from normalizer import normalize_for_matching, fuzzy_match
from excel_parser import parse_master_xlsx
from database import DatabaseManager
from config import AppConfig


def _generate_xlsx(path: Path, n: int) -> None:
    """Generate a synthetic master .xlsx with *n* study rows."""
    wb = Workbook()
    ws = wb.active

    phases = ["Phase I", "Phase II–IV"]
    subcats = ["Oncology", "Cardiology", "Neurology", "Immunology", "Dermatology"]
    sponsors = ["Pfizer", "Novartis", "Roche", "BMS", "Merck", "AstraZeneca", "Bayer", "Biogen"]

    row = 1
    studies_written = 0
    phase_idx = 0
    subcat_idx = 0

    while studies_written < n:
        # Write phase heading every ~50 studies
        if studies_written % 50 == 0:
            ws.cell(row=row, column=1, value=phases[phase_idx % len(phases)])
            row += 1
            phase_idx += 1

        # Write subcategory every ~10 studies
        if studies_written % 10 == 0:
            ws.cell(row=row, column=1, value=subcats[subcat_idx % len(subcats)])
            row += 1
            subcat_idx += 1

        year = 2020 + (studies_written % 5)
        sponsor = sponsors[studies_written % len(sponsors)]
        protocol = f"PROTO-{studies_written:05d}"
        desc = f"A Phase {1 + studies_written % 3} study of {protocol} (treatment-{studies_written}) in patients with condition-{studies_written % 20}"
        masked = f"A Phase {1 + studies_written % 3} study of XXX in patients with condition-{studies_written % 20}"

        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2, value=f"{sponsor} {protocol}: {desc}")
        ws.cell(row=row, column=3, value=f"{sponsor}: {masked}")
        row += 1
        studies_written += 1

    wb.save(path)
    wb.close()


def _make_studies(n: int) -> list:
    """Generate n synthetic Study objects in memory."""
    sponsors = ["Pfizer", "Novartis", "Roche", "BMS", "Merck", "AstraZeneca"]
    studies = []
    for i in range(n):
        studies.append(
            Study(
                phase="Phase I" if i % 2 == 0 else "Phase II–IV",
                subcategory=f"Subcat-{i % 10}",
                year=2020 + (i % 5),
                sponsor=sponsors[i % len(sponsors)],
                protocol=f"PROTO-{i:05d}",
                description_full=f"Study {i} full description with treatment-{i}",
                description_masked=f"Study {i} masked description with XXX",
            )
        )
    return studies


def bench_parse(n: int) -> float:
    """Benchmark parsing an xlsx with n studies."""
    with tempfile.TemporaryDirectory() as td:
        xlsx_path = Path(td) / "bench.xlsx"
        _generate_xlsx(xlsx_path, n)

        t0 = time.perf_counter()
        studies = parse_master_xlsx(xlsx_path)
        elapsed = time.perf_counter() - t0

    print(f"  Parse {n} studies from xlsx: {elapsed:.3f}s  ({len(studies)} parsed)")
    return elapsed


def bench_normalize(n: int) -> float:
    """Benchmark normalisation of n strings."""
    texts = [
        f"Pfizer PF-{i:05d}: A Phase 1 study of PF-{i:05d} in patients"
        for i in range(n)
    ]
    t0 = time.perf_counter()
    for text in texts:
        normalize_for_matching(text)
    elapsed = time.perf_counter() - t0
    print(f"  Normalize {n} strings:       {elapsed:.3f}s")
    return elapsed


def bench_fuzzy(n: int) -> float:
    """Benchmark fuzzy matching: n queries against a pool of 100 candidates."""
    pool = [
        f"Sponsor-{i} PROTO-{i:03d}: description for study number {i}"
        for i in range(100)
    ]
    queries = [
        f"Sponsor-{i % 100} PROTO-{i % 100:03d}: description for study number {i % 100}"
        for i in range(n)
    ]

    t0 = time.perf_counter()
    matches = 0
    for q in queries:
        for p in pool:
            is_match, _score = fuzzy_match(q, p, threshold=90)
            if is_match:
                matches += 1
                break
    elapsed = time.perf_counter() - t0
    print(f"  Fuzzy match {n} x 100 pool:  {elapsed:.3f}s  ({matches} matched)")
    return elapsed


def bench_db_insert(n: int) -> float:
    """Benchmark bulk SQLite insert of n studies."""
    studies = _make_studies(n)
    with tempfile.TemporaryDirectory() as td:
        config = AppConfig(data_root=td)
        config.ensure_user_directories()

        with DatabaseManager(config=config) as db:
            site = db.create_site("BenchSite")

            t0 = time.perf_counter()
            count = db.bulk_add_studies(site.id, studies)
            elapsed = time.perf_counter() - t0

        print(f"  DB bulk insert {n} studies: {elapsed:.3f}s  ({count} inserted)")
    return elapsed


def bench_db_read(n: int) -> float:
    """Benchmark reading n studies from SQLite."""
    studies = _make_studies(n)
    with tempfile.TemporaryDirectory() as td:
        config = AppConfig(data_root=td)
        config.ensure_user_directories()

        with DatabaseManager(config=config) as db:
            site = db.create_site("BenchSite")
            db.bulk_add_studies(site.id, studies)

            t0 = time.perf_counter()
            loaded = db.get_studies(site.id)
            elapsed = time.perf_counter() - t0

        print(f"  DB read {n} studies:        {elapsed:.3f}s  ({len(loaded)} read)")
    return elapsed


def main():
    parser = argparse.ArgumentParser(description="CV Manager Benchmark")
    parser.add_argument(
        "--count", "-n",
        type=int,
        default=1000,
        help="Number of studies to benchmark (default: 1000)",
    )
    args = parser.parse_args()
    n = args.count

    print(f"\n{'='*60}")
    print(f"  CV Manager Benchmark — {n} studies")
    print(f"{'='*60}\n")

    results = {}
    results["parse"] = bench_parse(n)
    results["normalize"] = bench_normalize(n)
    results["fuzzy"] = bench_fuzzy(min(n, 500))  # Cap fuzzy to avoid O(n*m)
    results["db_insert"] = bench_db_insert(n)
    results["db_read"] = bench_db_read(n)

    print(f"\n{'='*60}")
    print("  Summary")
    print(f"{'='*60}")
    for name, elapsed in results.items():
        print(f"  {name:20s}: {elapsed:.3f}s")
    print()


if __name__ == "__main__":
    main()
